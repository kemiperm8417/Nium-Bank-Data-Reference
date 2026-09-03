"""
Bank Reference Data — Nium Reference Data API → multi-sheet Excel.

Fetches bank and branch reference data per country from the public Nium
Reference Data service and normalises each country's country-specific JSON
into one flat row shape carrying the *domestic routing code* — ACH/ABA for
US, IFSC for IN, Sort Code for GB, BSB for AU, bank + branch code for JP/CA,
and BIC/SWIFT for SEPA countries that route on IBAN instead.

No credentials are required; every call is a read-only GET.

Library:
    rows = fetch_country("GB").rows
    xlsx = build_workbook([fetch_country("GB")])

CLI:
    python3 refdata.py --countries US,GB,AU --out banks.xlsx
    python3 refdata.py --sepa --out sepa.xlsx
    python3 refdata.py --probe IN
"""

from __future__ import annotations

import argparse
import gzip
import http.client
import io
import json
import logging
import random
import re
import socket
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import date
from typing import Any, Callable, Dict, List, Optional, Sequence, Set, Tuple

log = logging.getLogger(__name__)

import os
BASE_URL = os.environ.get("REFDATA_BASE_URL", "https://refdata.prod.nium.com/ref-data-service")

# True when running in the browser under Pyodide (stlite/GitHub Pages): no
# threads, and the browser owns HTTP (CORS, gzip) instead of urllib sockets.
IN_BROWSER = sys.platform == "emscripten"
CLIENT_ID = "nium-refdata-explorer"

# branchCodes silently truncates to 100 rows unless an explicit limit is sent.
BIG_LIMIT = 1_000_000

HTTP_TIMEOUT = 180          # IN per-bank chunks run ~11s; whole countries ~40s
MAX_RETRIES = 3
CHUNK_WORKERS = 6           # unauthenticated public endpoint — stay polite

EXCEL_MAX_ROWS = 1_048_575  # 1,048,576 minus the header row
EXCEL_MAX_CELL = 32_767

ProgressCb = Optional[Callable[[str, float], None]]


class RefDataError(RuntimeError):
    """A Reference Data API call failed after exhausting retries."""


class RefDataHTTPError(RefDataError):
    """Non-retryable HTTP status, or a 5xx that survived every retry."""

    def __init__(self, message: str, status: Optional[int] = None):
        super().__init__(message)
        self.status = status


# ─────────────────────────────────────────────────────────────────────────────
# HTTP
# ─────────────────────────────────────────────────────────────────────────────
def _get_json(path: str,
              params: Optional[Dict[str, Any]] = None,
              timeout: int = HTTP_TIMEOUT,
              retries: int = MAX_RETRIES) -> Any:
    """GET ``{BASE_URL}{path}`` and return the decoded JSON body.

    Sends ``X-CLIENT-ID`` (the spec declares it required; the service does not
    enforce it) and ``Accept-Encoding: gzip`` — gzip cuts GB/branchCodes from
    8.6 MB to 0.9 MB and halves wall-clock. ``urllib`` does not decompress
    automatically, so the body is inflated here when the response says so.

    Retries 5xx / 429 / timeouts / truncated reads with exponential backoff.
    Any other 4xx raises immediately.

    Raises:
        RefDataHTTPError: on a non-retryable status or a 5xx that never cleared.
        RefDataError: on repeated transport or decode failures.
    """
    url = BASE_URL + path
    if params:
        url += "?" + urllib.parse.urlencode(params)

    last: Optional[Exception] = None

    for attempt in range(1, retries + 1):
        try:
            headers = {"X-CLIENT-ID": CLIENT_ID, "Accept": "application/json"}
            if not IN_BROWSER:
                # Browsers forbid setting these on fetch and handle gzip themselves.
                headers["Accept-Encoding"] = "gzip"
                headers["User-Agent"] = "nium-refdata-explorer/1.0"
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                raw = resp.read()
                if (resp.headers.get("Content-Encoding") or "").lower() == "gzip":
                    try:
                        raw = gzip.decompress(raw)
                    except (OSError, EOFError):
                        pass    # browser fetch already inflated it
            return json.loads(raw)

        except urllib.error.HTTPError as exc:
            last = exc
            if exc.code < 500 and exc.code != 429:
                raise RefDataHTTPError("%s → HTTP %s" % (path, exc.code), exc.code) from exc
            log.warning("%s → HTTP %s (attempt %d/%d)", path, exc.code, attempt, retries)

        except (urllib.error.URLError, socket.timeout, TimeoutError,
                http.client.IncompleteRead, ConnectionResetError,
                json.JSONDecodeError) as exc:
            last = exc
            log.warning("%s → %s (attempt %d/%d)", path, exc, attempt, retries)

        except Exception as exc:                    # noqa: BLE001
            # In the browser (Pyodide) a blocked request surfaces as a
            # JavaScript exception (pyodide.ffi.JsException: NetworkError…),
            # not a urllib error. Almost always CORS, or not being on the VPN.
            if not IN_BROWSER:
                raise
            raise RefDataError(
                "%s blocked by the browser (%s). Either this page's origin is not "
                "in the API's CORS allow-list, or this machine is not on the Nium "
                "VPN." % (path, exc.__class__.__name__)) from exc

        if attempt < retries and not IN_BROWSER:      # time.sleep is a no-op in Pyodide
            time.sleep(min(30.0, 2.0 ** attempt) + random.uniform(0, 1))

    status = getattr(last, "code", None)
    raise RefDataHTTPError(
        "%s failed after %d attempt(s): %s" % (path, retries, last), status)


# ─────────────────────────────────────────────────────────────────────────────
# Routing codes
# ─────────────────────────────────────────────────────────────────────────────
@dataclass(frozen=True)
class RoutingSpec:
    """How one country's domestic routing code is spelled in the API payload."""

    label: str                          # e.g. "IFSC"
    branch_field: str = "branchId"      # branch key holding the routing code
    bank_field: Optional[str] = None    # bank key holding a bank/institution code
    bank_label: Optional[str] = None    # e.g. "Institution Number"


# Verified empirically against /{country}/branchCodes.
ROUTING_CODES: Dict[str, RoutingSpec] = {
    "US": RoutingSpec("ACH / ABA Routing Number", bank_field="abaCode", bank_label="ABA"),
    "IN": RoutingSpec("IFSC", bank_field="bankId", bank_label="Bank Code"),
    "GB": RoutingSpec("Sort Code"),
    "AU": RoutingSpec("BSB", branch_field="bsbCode"),
    "CA": RoutingSpec("Transit Number", bank_field="bankId", bank_label="Institution Number"),
    "JP": RoutingSpec("Branch Code", bank_field="bankId", bank_label="Bank Code"),
    "NZ": RoutingSpec("Branch Code", bank_field="bankId", bank_label="Bank Code"),
    "HK": RoutingSpec("Branch Code", bank_field="bankId", bank_label="Bank Code"),
    "ZA": RoutingSpec("Branch Code", bank_field="bankId", bank_label="Bank Code"),
    "TH": RoutingSpec("Branch Code", bank_field="bankId", bank_label="Bank Code"),
    "ID": RoutingSpec("Branch Code", bank_field="clearingCode", bank_label="Clearing Code"),
    "VN": RoutingSpec("Branch Code", bank_field="bankId", bank_label="Bank Code"),
    "PL": RoutingSpec("Branch Code", bank_field="bankId", bank_label="Bank Code"),
}

# Presence-driven fallback for countries with no explicit spec, so a country
# Nium adds later still exports a usable code instead of a blank column.
_GENERIC_FIELDS: Sequence[Tuple[str, str]] = (
    ("bsbCode", "BSB"),
    ("clearingCode", "Clearing Code"),
    ("abaCode", "ACH / ABA Routing Number"),
    ("branchCode", "Branch Code"),
    ("branchId", "Branch / Routing Code"),
)

# SEPA scheme membership: EU27 + IS LI NO CH GB MC SM AD VA.
# There is NO API endpoint for this — hand-maintained; review if SEPA changes.
# NOTE: GB is a SEPA member but is fetched from branchCodes (real sort codes),
# not swiftCodes. SEPA membership drives the preset button, not the strategy.
SEPA_COUNTRIES: Tuple[str, ...] = (
    "AT", "BE", "BG", "HR", "CY", "CZ", "DK", "EE", "FI", "FR", "DE", "GR",
    "HU", "IE", "IT", "LV", "LT", "LU", "MT", "NL", "PL", "PT", "RO", "SK",
    "SI", "ES", "SE",                                       # EU27
    "IS", "LI", "NO", "CH", "GB", "MC", "SM", "AD", "VA",   # +9
)

COMMON_CORRIDORS: Tuple[str, ...] = (
    "US", "GB", "AU", "CA", "IN", "SG", "HK", "NZ", "JP",
    "MX", "BR", "ZA", "AE", "PH", "ID", "MY", "TH", "VN",
)

# /IN/branchCodes returns HTTP 500 on an unbounded call — must chunk per bank.
CHUNK_PER_BANK: Set[str] = {"IN"}

# Countries verified to return zero branch records — bank-level only.
BANK_LEVEL_ONLY: Set[str] = {"MX", "BR", "SG", "AE", "PH", "MY", "KR", "CN", "TR"}

COLUMNS: List[str] = [
    "Country", "Country Name", "Bank Name", "Bank Code", "Routing Code Type",
    "Routing Code", "Branch Name", "Branch ID", "BIC / SWIFT", "BIC Source",
    "City", "State / Province", "Address", "Postal Code", "Source",
]

_BANK_CONSUMED = {
    "id", "label", "bankName", "bankId", "abaCode", "clearingCode", "bicCode",
    "sepaBankName", "refDataBankUniqueId",
}
_BRANCH_CONSUMED = {
    "id", "label", "bankName", "bankId", "branchName", "branchId", "branchCode",
    "bicCode", "branchBicCode", "bsbCode", "city", "state", "province",
    "proviceCode", "district", "address", "physicalAddress1", "physicalAddress2",
    "postalCode", "institutionName", "refDataBankBranchUniqueId",
}

_ILLEGAL_XLSX = re.compile(r"[\000-\010\013\014\016-\037]")


def _s(value: Any) -> str:
    """Coerce an API value to a trimmed, Excel-safe string; None becomes ''."""
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        value = json.dumps(value, separators=(",", ":"))
    text = str(value).strip()
    text = _ILLEGAL_XLSX.sub("", text)
    return text[:EXCEL_MAX_CELL]


def _first(*values: Any) -> str:
    """Return the first non-empty value, stringified."""
    for v in values:
        s = _s(v)
        if s:
            return s
    return ""


def _resolve_routing(country: str, bank: dict, branch: dict, bic: str = "") -> Tuple[str, str]:
    """Return ``(routing_code_type, routing_code)`` for one bank/branch pair.

    Uses :data:`ROUTING_CODES` when the country is mapped, otherwise infers the
    type from the fields the API actually returned for this record.
    """
    spec = ROUTING_CODES.get(country)
    if spec:
        val = _s(branch.get(spec.branch_field))
        if val:
            return spec.label, val

    for key, label in _GENERIC_FIELDS:
        val = _first(branch.get(key), bank.get(key))
        if val:
            return label, val

    code = _first(bic, branch.get("bicCode"), branch.get("branchBicCode"), bank.get("bicCode"))
    if code:
        return "BIC / SWIFT", code

    bank_id = _s(bank.get("bankId"))
    if bank_id:
        return "Bank Code", bank_id
    return "", ""


def _row(country: str, country_name: str, bank: dict, branch: dict,
         source: str, bic: str = "", extras: Optional[Set[str]] = None) -> Dict[str, str]:
    """Flatten one bank/branch pair into the canonical :data:`COLUMNS` shape.

    Any country-specific key the API returned that is not already consumed is
    appended as ``"Bank: <key>"`` / ``"Branch: <key>"`` so nothing is lost
    (IN ``upiEnabled``, AU ``nppEnabled``, US ``fedwireEnable``, DE ``systemType``…).
    """
    spec = ROUTING_CODES.get(country)
    rtype, rcode = _resolve_routing(country, bank, branch, bic)

    bank_code = ""
    if spec and spec.bank_field:
        bank_code = _first(bank.get(spec.bank_field), branch.get(spec.bank_field))
    if not bank_code:
        # GB/AU/HK/NZ return bankId as an empty string — fall through.
        bank_code = _first(bank.get("bankId"), bank.get("abaCode"),
                           bank.get("clearingCode"), bank.get("refDataBankUniqueId"))

    address = _first(
        branch.get("address"),
        ", ".join(p for p in (_s(branch.get("physicalAddress1")),
                              _s(branch.get("physicalAddress2"))) if p),
    )

    # Track where the BIC came from so a directly-supplied code is
    # distinguishable from one matched in by bank name.
    bic_value, bic_source = "", ""
    for candidate, origin in ((bic, "swiftCodes"),
                              (branch.get("bicCode"), "branch"),
                              (branch.get("branchBicCode"), "branch"),
                              (bank.get("bicCode"), "bank")):
        val = _s(candidate)
        if val:
            bic_value, bic_source = val, origin
            break

    row = {
        "Country":           country,
        "Country Name":      country_name,
        "Bank Name":         _first(bank.get("bankName"), branch.get("bankName"),
                                    bank.get("sepaBankName")),
        "Bank Code":         bank_code,
        "Routing Code Type": rtype,
        "Routing Code":      rcode,
        "Branch Name":       _first(branch.get("branchName"), branch.get("institutionName")),
        "Branch ID":         _first(branch.get("branchId"), branch.get("branchCode")),
        "BIC / SWIFT":       bic_value,
        "BIC Source":        bic_source,
        "City":              _s(branch.get("city")),
        "State / Province":  _first(branch.get("state"), branch.get("province"),
                                    branch.get("proviceCode"), branch.get("district")),
        "Address":           address,
        "Postal Code":       _s(branch.get("postalCode")),
        "Source":            source,
    }

    for key, val in (bank or {}).items():
        if key not in _BANK_CONSUMED:
            col = "Bank: %s" % key
            row[col] = _s(val)
            if extras is not None:
                extras.add(col)
    for key, val in (branch or {}).items():
        if key not in _BRANCH_CONSUMED:
            col = "Branch: %s" % key
            row[col] = _s(val)
            if extras is not None:
                extras.add(col)
    return row


# ─────────────────────────────────────────────────────────────────────────────
# Country list
# ─────────────────────────────────────────────────────────────────────────────
_country_names: Dict[str, str] = {}


def list_countries() -> List[Dict[str, Any]]:
    """Return ``[{'code','name','banned'}, …]`` sorted by name.

    Sourced from ``GET /entity/Country``. Banned countries are kept but flagged
    so a picker can grey them out rather than hide them.
    """
    # Fail fast: this is the first call the UI makes, and a hung connection
    # (e.g. from a host that cannot reach Nium's network) must surface as an
    # error in seconds, not after minutes of retries.
    payload = _get_json("/entity/Country", {"limit": 1000}, timeout=15, retries=1)
    out: List[Dict[str, Any]] = []
    for c in payload or []:
        code = _first(c.get("code_2"), c.get("id"))
        name = _s(c.get("name")) or code
        if not code:
            continue
        _country_names[code] = name
        out.append({"code": code, "name": name,
                    "banned": _s(c.get("is_banned")).upper() == "Y"})
    out.sort(key=lambda c: c["name"])
    return out


def country_name(code: str) -> str:
    """Best-effort display name for an ISO-2 code (falls back to the code)."""
    if not _country_names:
        try:
            list_countries()
        except RefDataError as exc:
            log.warning("Could not load country names: %s", exc)
    return _country_names.get(code, code)


# ─────────────────────────────────────────────────────────────────────────────
# Fetch strategies
# ─────────────────────────────────────────────────────────────────────────────
def _flatten_pairs(country: str, name: str, payload: Any,
                   source: str, extras: Set[str]) -> List[Dict[str, str]]:
    """Flatten a ``[{"bank":[…], "branch":[…]}, …]`` payload into rows."""
    rows: List[Dict[str, str]] = []
    for rec in payload or []:
        if not isinstance(rec, dict):
            continue
        banks = rec.get("bank") or [{}]
        bank = banks[0] if banks else {}
        for branch in (rec.get("branch") or [{}]):
            rows.append(_row(country, name, bank, branch, source, extras=extras))
    return rows


def _fetch_branch(country: str, name: str, extras: Set[str],
                  retries: int = MAX_RETRIES) -> List[Dict[str, str]]:
    """Whole-country ``/{c}/branchCodes``. The explicit limit is mandatory."""
    payload = _get_json("/%s/branchCodes" % country, {"limit": BIG_LIMIT}, retries=retries)
    return _flatten_pairs(country, name, payload, "branchCodes", extras)


def _fetch_per_bank(country: str, name: str, extras: Set[str],
                    progress_cb: ProgressCb = None) -> Tuple[List[Dict[str, str]], str]:
    """Chunked ``/{c}/{bankId}/branchCodes``, one request per bank.

    Required for India, whose unbounded whole-country call 500s after ~59s.
    Requests run in a small thread pool; ``progress_cb`` is invoked only from
    this (the calling) thread, so it is safe to drive Streamlit widgets with.
    A bank that fails is logged and skipped rather than failing the country.
    """
    banks = _get_json("/%s/bankCodes" % country, timeout=90)
    bank_ids = [b.get("bankId") for b in (banks or []) if _s(b.get("bankId"))]
    if not bank_ids:
        raise RefDataError("%s: no bankIds available to chunk by" % country)

    rows: List[Dict[str, str]] = []
    skipped = 0
    total = len(bank_ids)

    def _one(bid: str) -> Any:
        # bankIds can contain spaces (VN: "STANDARD CHARTERED") — encode the segment.
        return _get_json("/%s/%s/branchCodes" % (country, urllib.parse.quote(bid, safe="")),
                         {"limit": BIG_LIMIT})

    if IN_BROWSER:
        # Pyodide has no threads — fetch banks one after another.
        pending = ((bid, None) for bid in bank_ids)
    else:
        pool = ThreadPoolExecutor(max_workers=CHUNK_WORKERS)
        futures = {pool.submit(_one, bid): bid for bid in bank_ids}
        pending = ((futures[f], f) for f in as_completed(futures))

    for i, (bid, fut) in enumerate(pending, start=1):
        try:
            payload = fut.result() if fut is not None else _one(bid)
            rows.extend(_flatten_pairs(country, name, payload,
                                       "branchCodes:per-bank", extras))
        except Exception as exc:                    # noqa: BLE001 - skip, never abort
            skipped += 1
            log.warning("%s/%s branches failed: %s", country, bid, exc)
        if progress_cb:
            progress_cb("%s — bank %d/%d (%s)" % (country, i, total, bid), i / total)

    if not IN_BROWSER:
        pool.shutdown(wait=True)

    note = "chunked over %d banks" % total
    if skipped:
        note += "; %d bank(s) skipped" % skipped
    return rows, note


def _fetch_swift(country: str, name: str, extras: Set[str]) -> List[Dict[str, str]]:
    """``/{c}/swiftCodes`` — the SEPA-sourced directory. One row per BIC."""
    payload = _get_json("/%s/swiftCodes" % country, {"limit": BIG_LIMIT})
    rows: List[Dict[str, str]] = []
    for rec in payload or []:
        if not isinstance(rec, dict):
            continue
        banks = rec.get("bank") or [{}]
        bank = banks[0] if banks else {}
        bics = rec.get("bics") or []
        if not bics:
            rows.append(_row(country, name, bank, {}, "swiftCodes", extras=extras))
            continue
        for bic in bics:
            rows.append(_row(country, name, bank, {}, "swiftCodes",
                             bic=_s(bic.get("bicCode")), extras=extras))
    return rows


def _fetch_bank_level(country: str, name: str, extras: Set[str]) -> List[Dict[str, str]]:
    """Last resort: ``/{c}/bankCodes`` merged with any BICs from swiftCodes."""
    banks = _get_json("/%s/bankCodes" % country, timeout=120)

    bic_by_name: Dict[str, str] = {}
    try:
        for rec in _get_json("/%s/swiftCodes" % country, {"limit": BIG_LIMIT}) or []:
            b = (rec.get("bank") or [{}])[0]
            key = _s(b.get("bankName")).upper()
            for bic in (rec.get("bics") or []):
                if key and key not in bic_by_name:
                    bic_by_name[key] = _s(bic.get("bicCode"))
    except RefDataError as exc:
        log.info("%s: no swiftCodes to merge (%s)", country, exc)

    rows = []
    for bank in banks or []:
        bic = _s(bank.get("bicCode")) or bic_by_name.get(_s(bank.get("bankName")).upper(), "")
        rows.append(_row(country, name, bank, {}, "bankCodes", bic=bic, extras=extras))
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# SWIFT/BIC enrichment for domestic-routing countries
# ─────────────────────────────────────────────────────────────────────────────
# Legal-form suffixes stripped for the loose match tier. Kept deliberately
# short — over-aggressive stripping collapses distinct banks onto one key.
_BANK_SUFFIXES = re.compile(
    r"\b(LIMITED|LTD|PLC|INC|INCORPORATED|CORPORATION|CORP|COMPANY|"
    r"NATIONAL ASSOCIATION|NA|NV|SA|AG|GMBH|PTE|PVT|PUBLIC|JOINT STOCK)\b")


def _norm_bank(name: Any) -> str:
    """Uppercase, drop punctuation, collapse whitespace."""
    s = re.sub(r"[^A-Z0-9 ]+", " ", _s(name).upper())
    return re.sub(r"\s+", " ", s).strip()


def _norm_bank_loose(name: Any) -> str:
    """:func:`_norm_bank` with parentheticals and legal-form suffixes removed.

    ``"HSBC BANK PLC (RFB)"`` → ``"HSBC BANK"``.
    """
    bare = re.sub(r"\([^)]*\)", " ", _s(name))
    return re.sub(r"\s+", " ", _BANK_SUFFIXES.sub(" ", _norm_bank(bare))).strip()


def _bic_index(country: str) -> Tuple[Dict[str, List[str]], Dict[str, List[str]]]:
    """Build ``(exact, loose)`` bank-name → sorted distinct BICs from ``/{c}/swiftCodes``.

    Large banks legitimately carry several BICs (Barclays has 2, HSBC 12,
    Australia's Cuscal 50+ as a credit-union aggregator), so every distinct
    code is kept — the caller decides how to present a multi-BIC bank.
    """
    try:
        payload = _get_json("/%s/swiftCodes" % country, {"limit": BIG_LIMIT},
                            timeout=120, retries=2)
    except RefDataError as exc:
        log.info("%s: no swiftCodes to merge (%s)", country, exc)
        return {}, {}

    exact: Dict[str, Set[str]] = {}
    loose: Dict[str, Set[str]] = {}
    for rec in payload or []:
        if not isinstance(rec, dict):
            continue
        bank = (rec.get("bank") or [{}])[0]
        name = _first(bank.get("bankName"), bank.get("sepaBankName"))
        if not name:
            continue
        for bic in (rec.get("bics") or []):
            code = _s(bic.get("bicCode"))
            if not code:
                continue
            exact.setdefault(_norm_bank(name), set()).add(code)
            loose.setdefault(_norm_bank_loose(name), set()).add(code)

    return ({k: sorted(v) for k, v in exact.items() if k},
            {k: sorted(v) for k, v in loose.items() if k})


def _enrich_bics(country: str, rows: List[Dict[str, str]]) -> str:
    """Fill empty ``BIC / SWIFT`` cells by matching bank name against swiftCodes.

    Only for countries that already carry a *domestic* routing code — the point
    is to show the SWIFT code alongside the BSB/IFSC/sort code, not to replace
    it. Rows that already have a BIC from the payload are left untouched.
    """
    blanks = [r for r in rows if not r.get("BIC / SWIFT")]
    if not blanks:
        return ""

    exact, loose = _bic_index(country)
    if not exact and not loose:
        return "no SWIFT data to merge"

    # Resolve once per distinct bank name, not once per branch row.
    resolved: Dict[str, Tuple[str, str]] = {}
    for row in blanks:
        name = row.get("Bank Name") or ""
        if name not in resolved:
            codes = exact.get(_norm_bank(name))
            origin = "swiftCodes (name)"
            if not codes:
                codes = loose.get(_norm_bank_loose(name))
                origin = "swiftCodes (loose)"
            if not codes:
                resolved[name] = ("", "")
            elif len(codes) == 1:
                resolved[name] = (codes[0], origin)
            else:
                resolved[name] = (" / ".join(codes), "swiftCodes (multiple)")
        code, origin = resolved[name]
        if code:
            row["BIC / SWIFT"] = code
            row["BIC Source"] = origin

    filled = sum(1 for r in blanks if r.get("BIC / SWIFT"))
    matched_banks = sum(1 for v in resolved.values() if v[0])
    return "SWIFT merged into %s of %s blank rows (%s/%s banks matched)" % (
        format(filled, ","), format(len(blanks), ","),
        matched_banks, len(resolved))


_mode_cache: Dict[str, str] = {}


def choose_mode(country: str) -> str:
    """Pick the endpoint strategy for a country: ``branch``/``branch_chunked``/``swift``.

    1. India and anything else known to 500 unbounded → ``branch_chunked``.
    2. A country with an explicit :data:`ROUTING_CODES` entry has real domestic
       branch data → ``branch``. This is checked BEFORE SEPA so GB keeps its
       20,567 sort codes instead of degrading to BICs.
    3. Known zero-branch countries and remaining SEPA members → ``swift``.
    4. Otherwise probe ``/{c}/branchCodes`` with no limit — the server's own
       100-row default makes the probe cheap. 5+ rows → ``branch``, else ``swift``.
    """
    if country in _mode_cache:
        return _mode_cache[country]

    if country in CHUNK_PER_BANK:
        mode = "branch_chunked"
    elif country in ROUTING_CODES:
        mode = "branch"
    elif country in BANK_LEVEL_ONLY or country in SEPA_COUNTRIES:
        mode = "swift"
    else:
        try:
            probe = _get_json("/%s/branchCodes" % country, timeout=60, retries=1)
            mode = "branch" if len(probe or []) >= 5 else "swift"
        except RefDataError as exc:
            log.info("%s: branch probe failed (%s) — falling back to swift", country, exc)
            mode = "swift"

    _mode_cache[country] = mode
    return mode


@dataclass
class CountryResult:
    """Outcome of fetching one country — rows plus enough context for a summary."""

    code: str
    name: str
    mode: str
    rows: List[Dict[str, str]] = field(default_factory=list)
    extra_keys: List[str] = field(default_factory=list)
    seconds: float = 0.0
    error: Optional[str] = None
    notes: str = ""


def fetch_country(code: str, progress_cb: ProgressCb = None,
                  mode: Optional[str] = None) -> CountryResult:
    """Fetch and normalise every bank/branch record for one ISO-2 country.

    Degrades automatically between strategies — a 5xx on the whole-country call
    retries as per-bank chunks, and an empty branch result falls back to
    bank-level data — so a country yields *something* wherever possible.
    Failures are captured on the result rather than raised, so one bad country
    never aborts a multi-country export.
    """
    code = code.strip().upper()
    name = country_name(code)
    started = time.time()
    mode = mode or choose_mode(code)
    extras: Set[str] = set()
    rows: List[Dict[str, str]] = []
    notes = ""

    try:
        if mode == "branch_chunked":
            rows, notes = _fetch_per_bank(code, name, extras, progress_cb)

        elif mode == "branch":
            if progress_cb:
                progress_cb("%s — fetching branches…" % code, 0.1)
            try:
                rows = _fetch_branch(code, name, extras)
            except RefDataError as exc:
                # The India failure mode, arriving in some other country.
                log.warning("%s: whole-country branchCodes failed (%s) — chunking", code, exc)
                if progress_cb:
                    progress_cb("%s — too large, chunking per bank…" % code, 0.15)
                rows, notes = _fetch_per_bank(code, name, extras, progress_cb)
                mode = "branch_chunked"
                notes = ("degraded from branch; " + notes).strip()

        else:
            if progress_cb:
                progress_cb("%s — fetching SWIFT/SEPA directory…" % code, 0.1)
            rows = _fetch_swift(code, name, extras)

        if not rows:
            if progress_cb:
                progress_cb("%s — no branch data, using bank level…" % code, 0.6)
            rows = _fetch_bank_level(code, name, extras)
            mode = "bankCodes"
            notes = (notes + "; no branch data — bank level only").strip("; ")

        # Countries with a domestic routing code also get their SWIFT/BIC
        # merged in, so each row carries both. Skipped for swift-mode
        # countries, where the BIC already IS the routing code.
        if mode in ("branch", "branch_chunked") and rows:
            if progress_cb:
                progress_cb("%s — merging SWIFT codes…" % code, 0.95)
            merge_note = _enrich_bics(code, rows)
            if merge_note:
                notes = "; ".join(n for n in (notes, merge_note) if n)

    except RefDataError as exc:
        return CountryResult(code, name, mode, [], sorted(extras),
                             time.time() - started, str(exc), notes)
    except Exception as exc:                      # noqa: BLE001 - isolate the country
        log.exception("%s: unexpected failure", code)
        return CountryResult(code, name, mode, [], sorted(extras),
                             time.time() - started, repr(exc), notes)

    if progress_cb:
        progress_cb("%s — %s rows" % (code, format(len(rows), ",")), 1.0)

    return CountryResult(code, name, mode, rows, sorted(extras),
                         time.time() - started, None, notes)


def fetch_many(codes: Sequence[str], progress_cb: ProgressCb = None) -> List[CountryResult]:
    """Fetch countries in order, isolating failures.

    Countries run serially — chunking already uses a thread pool internally,
    and serial ordering keeps the progress display honest.
    """
    results: List[CountryResult] = []
    total = max(len(codes), 1)
    for i, code in enumerate(codes):
        def _cb(msg: str, frac: float, _i=i) -> None:
            if progress_cb:
                progress_cb(msg, min((_i + frac) / total, 1.0))
        results.append(fetch_country(code, _cb))
    return results


# ─────────────────────────────────────────────────────────────────────────────
# Excel
# ─────────────────────────────────────────────────────────────────────────────
_ILLEGAL_SHEET = re.compile(r"[\[\]:*?/\\]")


def _safe_sheet_name(code: str, name: str, used: Set[str]) -> str:
    """Build a unique, Excel-legal sheet name, e.g. ``"GB United Kingdom"``.

    Excel allows at most 31 characters and forbids ``[ ] : * ? / \\``. The ISO
    code is kept as the prefix so truncation never makes two sheets ambiguous,
    and the de-duplication suffix is applied *inside* the 31-char budget.
    """
    label = _ILLEGAL_SHEET.sub(" ", "%s %s" % (code, name)).strip()
    label = re.sub(r"\s+", " ", label).strip("'") or code
    if label.lower() == "history":          # reserved by Excel
        label = "%s country" % code
    candidate = label[:31]

    n = 2
    while candidate.lower() in used:
        suffix = " (%d)" % n
        candidate = label[:31 - len(suffix)] + suffix
        n += 1
    used.add(candidate.lower())
    return candidate


def _headers_for(result: "CountryResult") -> List[str]:
    """Canonical columns followed by this country's extra fields."""
    return COLUMNS + [k for k in result.extra_keys if k not in COLUMNS]


def build_workbook(results: Sequence["CountryResult"]) -> bytes:
    """Render fetched countries into a multi-sheet ``.xlsx`` and return its bytes.

    Sheet 1 is always ``Summary`` (per-country status, row count, endpoint used,
    elapsed time, notes and any error) so a partial export is self-documenting.
    Then one sheet per country.

    Uses openpyxl's *write-only* workbook and appends row by row, so India's
    ~170k branches never materialise as a full in-memory cell grid.
    """
    import openpyxl
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Font

    wb = openpyxl.Workbook(write_only=True)
    bold = Font(bold=True)

    def _write_header(ws, headers: List[str]) -> None:
        ws.freeze_panes = "A2"          # must be set before appending
        cells = []
        for h in headers:
            c = WriteOnlyCell(ws, value=h)
            c.font = bold
            cells.append(c)
        ws.append(cells)

    # ── Summary ──────────────────────────────────────────────────────────────
    summary = wb.create_sheet("Summary")
    _write_header(summary, ["Country", "Country Name", "Rows", "Routing Code Type",
                            "Endpoint", "Seconds", "Status", "Notes"])
    for r in results:
        rtype = ""
        for row in r.rows[:50]:
            if row.get("Routing Code Type"):
                rtype = row["Routing Code Type"]
                break
        summary.append([
            r.code, r.name, len(r.rows), rtype, r.mode, round(r.seconds, 1),
            "ERROR" if r.error else ("EMPTY" if not r.rows else "OK"),
            r.error or r.notes,
        ])

    # ── One sheet per country ────────────────────────────────────────────────
    used: Set[str] = {"summary"}
    for r in results:
        headers = _headers_for(r)
        chunks = [r.rows[i:i + EXCEL_MAX_ROWS]
                  for i in range(0, max(len(r.rows), 1), EXCEL_MAX_ROWS)] or [[]]
        for idx, chunk in enumerate(chunks):
            label = r.name if idx == 0 else "%s p%d" % (r.name, idx + 1)
            ws = wb.create_sheet(_safe_sheet_name(r.code, label, used))
            _write_header(ws, headers)
            for row in chunk:
                ws.append([row.get(h, "") for h in headers])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def suggested_filename(codes: Sequence[str]) -> str:
    """e.g. ``Bank_Reference_Data_US-GB-AU_20260903.xlsx``.

    Collapses to a count past four countries, and to ``SEPA36`` when the
    selection is exactly the SEPA membership list.
    """
    codes = [c.strip().upper() for c in codes if c.strip()]
    if set(codes) == set(SEPA_COUNTRIES):
        label = "SEPA36"
    elif not codes:
        label = "empty"
    elif len(codes) <= 4:
        label = "-".join(codes)
    else:
        label = "%d_countries" % len(codes)
    label = re.sub(r"[^A-Za-z0-9._-]+", "_", label).strip("_") or "export"
    return "Bank_Reference_Data_%s_%s.xlsx" % (label, date.today().strftime("%Y%m%d"))


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────
def _probe(code: str) -> None:
    """Print the strategy and observed field profile for one country.

    The maintenance tool: when Nium changes a payload shape, this shows the
    bank/branch keys actually returned so ROUTING_CODES can be updated.
    """
    code = code.upper()
    mode = choose_mode(code)
    print("%s → mode=%s" % (code, mode))
    try:
        payload = _get_json("/%s/branchCodes" % code, {"limit": 25}, timeout=60, retries=1)
        bank_keys: Set[str] = set()
        branch_keys: Set[str] = set()
        for rec in payload or []:
            for b in rec.get("bank", []):
                bank_keys |= set(b)
            for b in rec.get("branch", []):
                branch_keys |= set(b)
        print("  branchCodes probe: %d record(s)" % len(payload or []))
        print("    bank keys  : %s" % sorted(bank_keys))
        print("    branch keys: %s" % sorted(branch_keys))
    except RefDataError as exc:
        print("  branchCodes probe failed: %s" % exc)

    spec = ROUTING_CODES.get(code)
    print("  routing spec: %s" % (spec if spec else "(generic fallback)"))


def main(argv: Optional[Sequence[str]] = None) -> int:
    ap = argparse.ArgumentParser(
        description="Export Nium bank reference data to a multi-sheet Excel workbook.")
    ap.add_argument("--countries", help="Comma-separated ISO-2 codes, e.g. US,GB,AU")
    ap.add_argument("--sepa", action="store_true", help="Use the SEPA 36-member list")
    ap.add_argument("--common", action="store_true", help="Use the common-corridor list")
    ap.add_argument("--probe", help="Print the strategy and field profile for one country")
    ap.add_argument("--out", default="bank_reference_data.xlsx", help="Output .xlsx path")
    ap.add_argument("-v", "--verbose", action="store_true")
    args = ap.parse_args(argv)

    logging.basicConfig(
        level=logging.INFO if args.verbose else logging.WARNING,
        format="%(levelname)s %(name)s: %(message)s")

    if args.probe:
        _probe(args.probe)
        return 0

    if args.sepa:
        codes = list(SEPA_COUNTRIES)
    elif args.common:
        codes = list(COMMON_CORRIDORS)
    elif args.countries:
        codes = [c.strip().upper() for c in args.countries.split(",") if c.strip()]
    else:
        ap.error("one of --countries, --sepa, --common or --probe is required")

    def _cb(msg: str, frac: float) -> None:
        sys.stderr.write("\r%-72s %3.0f%%" % (msg[:72], frac * 100))
        sys.stderr.flush()

    started = time.time()
    results = fetch_many(codes, _cb)
    sys.stderr.write("\n")

    for r in results:
        status = "ERROR: %s" % r.error if r.error else "%s rows" % format(len(r.rows), ",")
        print("%-4s %-28s %-18s %8s  %5.1fs %s"
              % (r.code, r.name[:28], r.mode, status, r.seconds, r.notes))

    with open(args.out, "wb") as fh:
        fh.write(build_workbook(results))

    total = sum(len(r.rows) for r in results)
    print("\nWrote %s — %s rows across %d countries in %.1fs"
          % (args.out, format(total, ","), len(results), time.time() - started))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
